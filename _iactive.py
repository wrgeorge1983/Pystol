#! /usr/bin/python
"""
Created on Mar 26, 2015

@author: William.George

Credit to /r/Python for the non-wasteful and sensible handling of oldInit and
    newInit
"""
# Standard Library Imports
from collections import defaultdict
import collections
from datetime import date
import json
import multiprocessing.pool
import os
from pprint import pprint  # Not used here, but we want it in interactive mode.
from subprocess import Popen
import sys
import time

sys.path += [os.getcwd()]

# Imports from other modules in this project
import sshutil

# Imports from third party modules
# import phpipam
import ipaddress
import openpyxl

DEFAULT_SW_IP = '10.10.10.10'
DEFAULT_HOST_IP = '10.10.10.10'
DEFAULT_IPAM_HOST = 'ipam'
DEFAULT_IPAM_API_ID = 'ipam'
DEFAULT_IPAM_API_KEY = 'FFFFF'

try:
    import iactiveconstants

    DEFAULT_SW_IP = iactiveconstants.DEFAULT_SW_IP
    DEFAULT_HOST_IP = iactiveconstants.DEFAULT_HOST_IP
    DEFAULT_IPAM_HOST = iactiveconstants.DEFAULT_IPAM_HOST
    DEFAULT_IPAM_API_ID = iactiveconstants.DEFAULT_IPAM_API_ID
    DEFAULT_IPAM_API_KEY = iactiveconstants.DEFAULT_IPAM_API_KEY
except ImportError:
    pass

# File class from user fdb on StackOverflow
# http://stackoverflow.com/questions/5896079/python-head-tail-and-backward-read-by-lines-of-a-text-file
# class File(file):
#     """ An helper class for file reading  """
#
#     def __init__(self, *args, **kwargs):
#         super(File, self).__init__(*args, **kwargs)
#         self.BLOCKSIZE = 4096
#
#     def head(self, lines_2find=1):
#         self.seek(0)  # Rewind file
#         return [super(File, self).next() for x in xrange(lines_2find)]
#
#     def tail(self, lines_2find=1):
#         self.seek(0, 2)  # Go to end of file
#         bytes_in_file = self.tell()
#         lines_found, total_bytes_scanned = 0, 0
#         while (lines_2find + 1 > lines_found and
#                 bytes_in_file > total_bytes_scanned):
#             byte_block = min(
#                 self.BLOCKSIZE,
#                 bytes_in_file - total_bytes_scanned)
#             self.seek(-(byte_block + total_bytes_scanned), 2)
#             total_bytes_scanned += byte_block
#             lines_found += self.read(self.BLOCKSIZE).count('\n')
#         self.seek(-total_bytes_scanned, 2)
#         line_list = list(self.readlines())
#         return line_list[-lines_2find:]
#
#     def backward(self):
#         self.seek(0, 2)  # Go to end of file
#         blocksize = self.BLOCKSIZE
#         last_row = ''
#         while self.tell() != 0:
#             try:
#                 self.seek(-blocksize, 1)
#             except IOError:
#                 blocksize = self.tell()
#                 self.seek(-blocksize, 1)
#             block = self.read(blocksize)
#             self.seek(-blocksize, 1)
#             rows = block.split('\n')
#             rows[-1] = rows[-1] + last_row
#             while rows:
#                 last_row = rows.pop(-1)
#                 if rows and last_row:
#                     yield last_row
#         yield last_row

#
# def ipm(site, ipt):
#     """
#         ipm(site, ipt):
#             site: An IP or Network address in dotted-decimal in a string.
#                 e.g. "10.10.8.6" or "10.10.0.0"
#             ipt: 'input', trailing octets to be merged with site
#                 as string:
#                     e.g. "7" or "9.1.3"
#                 or as int or float:
#                     e.g. 7 or 3.8
#             Returns: trailing octets defined by ipt super-imposed on site
#                 e.g. site("10.10.8.6", "1.2.3") == "10.1.2.3"
#                      site("10.10.8.6", 5.1) == "10.10.5.1"
#             Note: It's possible that 'site' can be specified as 3 or fewer
#                 ('10.3', etc...) but this is probably not smart.
#             Note: This works exclusively by manipulating a string of octets
#                 in dotted decimal format, and does not in any way account for
#                 any real subnetting operations, etc...
#     """
#     ipt = str(ipt).split('.')
#     site = site.split('.')
#     return '.'.join(site[:4 - len(ipt)] + ipt)
#
#
# # TODO: This should really be wrapped in a class
# def pull_subnets():
#     ipam = phpipam.PHPIPAM(DEFAULT_IPAM_HOST, DEFAULT_IPAM_API_ID,
#                            DEFAULT_IPAM_API_KEY)
#     ipam.scheme = 'https'
#     rslt = ipam.read_subnets()
#     jload = json.loads(rslt)
#     subnets = jload['data']
#     return subnets
#
#
# def site_lookup(sfilter):
#     subnets = pull_subnets()
#     return [subnets[x] for x in range(0, len(subnets) - 1) if sfilter in
#             subnets[x]['description']]
#
#
# class IPAMController(object):
#     """Generic wrapper for JSON objects returned by ipam api"""
#
#     def __init__(self, ipam, data=None, **kwargs):
#         """Takes either the JSON data by itself or unpacked keywords.
#         if unpacked values are passed, ensure only the 'data' portion
#         of the result is sent.  i.e.:
#             rslt = ipam.read_subnets(id=1)
#             rslt = json.loads(rslt)['data']
#             subnet
#         """
#         self.ipam = ipam
#         if data is not None:
#             kwargs = json.loads(data)['data']
#
#         # Unsure if this is consistent or not, but I've seen it at least once
#         if type(kwargs) is list:
#             kwargs = kwargs[0]
#
#         for k, v in kwargs.items():
#             setattr(self, k, v)
#
#
# class IPAMSubnet(IPAMController):
#     """Wrap subnet JSON objects that come from phpipam"""
#
#     def __init__(self, **kwargs):
#         IPAMController.__init__(self, **kwargs)
#
#         net, mask = self.subnet, self.mask
#         try:
#             self.network = ipaddress.ip_network(u'{0}/{1}'.format(net, mask))
#         except ValueError:
#             self.network = 'INVALID'
#         self._site_codes = []
#
#     def _pull_site_codes(self):
#         subnet_id = self.id
#         addresses = self.ipam.generic('addresses', 'read', subnetId=subnet_id, format='ip')
#         addresses = json.loads(addresses)['data']
#         names = (x['dns_name'] for x in addresses)
#         site_codes = (x[5:8] for x in names)
#         self._site_codes = set(site_codes)
#
#     @property
#     def site_codes(self):
#         if len(self._site_codes) == 0:
#             self._pull_site_codes()
#
#         return self._site_codes
#
#     def __str__(self):
#         return str(self.network)
#
#
# class IPAM(phpipam.PHPIPAM):
#     """Handle subnets and addresses meaningfully"""
#
#     def __init__(self,
#                  url=DEFAULT_IPAM_HOST,
#                  api_id=DEFAULT_IPAM_API_ID,
#                  api_key=DEFAULT_IPAM_API_KEY,
#                  scheme='https'):
#         phpipam.PHPIPAM.__init__(self, url, api_id, api_key)
#         self.scheme = scheme
#         self._subnets = None
#         self._raw_subnets = None
#         self._addresse = None
#
#     def _pull_raw_subnets(self):
#         rslt = self.read_subnets()
#         jload = json.loads(rslt)
#         self._raw_subnets = jload['data']
#
#     @property
#     def raw_subnets(self):
#         if self._raw_subnets is None:
#             self._pull_raw_subnets()
#         return self._raw_subnets
#
#     def _pull_subnets(self):
#         self._subnets = {}
#         for subnet in self.raw_subnets:
#             self._subnets[subnet[u'id']] = IPAMSubnet(ipam=self, **subnet)
#
#     @property
#     def subnets(self, subnet_id=None):
#         """access one or all subnets"""
#         if self._subnets is None:
#             self._pull_subnets()
#
#         if subnet_id is not None:
#             return self._subnets[subnet_id]
#
#         return self._subnets
#
#     def audit_subnets(self):
#         rslt = True
#         for subnet in self.subnets.values():
#             try:
#                 net, mask = subnet.subnet, subnet.mask
#                 subnet.network = ipaddress.ip_network(u'{0}/{1}'.format(net, mask))
#             except ValueError as e:
#                 rslt = False
#                 print(e)
#         return rslt


# Wrapps Switch() with features that are great for interactive access,
# but would be terrible to use in an normal script.
class clintSwitch(sshutil.Switch):
    def __init__(self, ip=None, creds=None, timeout=None):
        if timeout:
            self.timeout = timeout
        elif not hasattr(self, 'timeout'):
            self.timeout = None

        if creds:
            clintSwitch.credentials = creds
        else:
            if not hasattr(self, "credentials"):
                raise SyntaxError("Credentials must be provided at least once.")
            creds = self.credentials
        if ip:
            ip = str(ip)
            site = ip
            ips = ip.split('.')
            if len(ips) == 4:
                clintSwitch.site = site
            else:
                if not hasattr(self, 'site'):
                    raise SyntaxError("Full IP must be provided at least once.")
                ip = ipm(clintSwitch.site, ip)
                clintSwitch.site = ip
        else:
            ip = 'None'
        sshutil.Switch.__init__(self, ip, creds)

    @property
    def flash_total(self):
        try:
            return self.flash.total
        except:
            return 'UNK'

    @property
    def flash_free(self):
        try:
            return self.flash.free
        except:
            return 'UNK'

    def pexecute(self, cmd, trim=True, timeout=None):
        args = [cmd, trim]
        if not timeout:
            timeout = self.timeout

        if timeout:
            args.append(timeout)

        print(self.execute(*args))

    def interact(self):
        cmd = 'ssh {0}'.format(self.ip)
        Popen(cmd, shell=True).communicate()

    def bufferflush(self):
        return self.connection.buffer_flush()


def poll_switch(sw, cmd, sleep_time):
    """sw.pexecute(cmd) every sleep_time seconds"""

    while True:
        sw.pexecute(cmd)
        time.sleep(sleep_time)


def pythonrc():
    """Return expanded path to current users .pythonrc.py"""
    home = os.path.expanduser('~/')
    return home + '.pythonrc.py'


def retrieve_pcaps(sw):
    destcreds = sshutil.get_credentials()
    host = DEFAULT_HOST_IP
    lines = sw.execute('sh flash: | i pcap').splitlines()
    files = [line.split()[-1] for line in lines]
    for fil in files:
        command = 'copy {0} scp:'.format(fil)
        sw.timeout = 2
        print('pulling {0}...'.format(fil))
        sw.pexecute(command)
        sw.pexecute(host)
        sw.pexecute('\n')
        sw.pexecute('\n')
        sw.pexecute(destcreds[1], 5)

class WorkbookWrapper(object):
    def __init__(self, filename):

        self.column_from_string = lambda x: openpyxl.utils.column_index_from_string(x) - 1

        self.wb = self.load_workbook(filename)
        self.ws = self.wb.active
        self.rows = self.ws.rows
        self.columns = self.ws.columns
        self.cell = self.ws.cell

        self.build_header()
        self.attribute_mapping = dict()

        self.attribute_mapping.update(  # maps header fields to object attributes
            {
                'hostname': 'hostname',
                'ip address': 'ip',
                'supervisor': 'supervisor',
                'ram (k)': 'available_ram',
                'total flash': 'flash_total',
                'free flash': 'flash_free',
                'model': 'model',
                'stacked': 'stacked',
                'old': 'software_version',
                'current': 'software_version',
                'feature set (license)': 'license'
            }
        )

    def build_header(self):
        """
        Assume header is row A
        :return:
        """
        header_row = self.rows[0]
        header = [(cell.value.lower(), index) for index, cell in enumerate(header_row)
                  if cell.value is not None]
        self.header = defaultdict(str)
        for (name, index) in header:
            self.header[name] = index
            self.header[index] = name

    def output_values(self, switches):
        """
        Takes switches (for now: manually provided, pre-populated) and outputs their attributes to xlsx.

        :param switches:
        :return:
        """
        am = self.attribute_mapping
        header = self.header

        for row, switch in zip(self.rows[1:], switches):  # skip header row obviously
            skipped = set()
            note = ''
            note_header = 'script notes'
            note_cell = row[int(header[note_header])]
            for index, cell in enumerate(row):
                if switch.state.upper() != 'UP':
                    note = 'unreachable as of {0}'.format(str(date.today()))
                    break

                header_text = header[index]
                try:
                    rslt = getattr(switch, str(am[header_text]), 'UNK')
                    if rslt == 'UNK':
                        raise AttributeError

                except (AttributeError, KeyError):
                    if header_text.strip().lower() != note_header:
                        skipped.add(header_text.lower())
                    elif skipped:
                        note = 'skipped: {0}'.format(
                            str(
                                skipped.intersection(
                                    self.attribute_mapping.keys()
                                )
                            )
                        )
                    continue

                cell.value = rslt
            note_cell.value = note



    # def validate_hostname(self, switch, value):
    #     if switch.hostname == value:
    #         return True, switch.hostname
    #     else:
    #         return False, switch.hostname
    #
    # def validate_supervisor(self, switch, value):
    #     sup = switch.supervisor
    #     return sup == value, sup
    #
    # @staticmethod
    # def validate_switch_attribute(switch, attribute, value):
    #     ref = getattr(switch, attribute)
    #     return ref == value, ref

    @staticmethod
    def load_workbook(filename):
        """
        return an xlsx document
        :param filename: filename of xlsx doc.  Assume it's under ~/stage/
        :return:
        """

        path = os.path.join('~', 'stage', filename)
        path = os.path.expanduser(path)
        wb = openpyxl.load_workbook(path)
        return wb

    def switch_from_row(self, row=None, row_index=None):
        if row is None:
            assert row_index is not None, "switch_from_row expects row or row_index"
            row = self.rows[row_index]

        assert row in self.rows, "row must be an existing row in rows"

        attrib_from_cell = lambda x: self.header[self.column_from_string(x.column)]
        attrs = dict((attrib_from_cell(cell), cell.value) for cell in row
                          if cell.value is not None)
        try:
            switch = clintSwitch(ip=attrs['ip address'])
        except KeyError:
            return None

        switch.row_index = row_index
        return switch

    def switches_from_rows(self):
        return [self.switch_from_row(row=row) for row in self.rows[1:]]  # skip header!

    def get_attribs(self, switch):
        pass


# TODO: These are here only for testing purposes and should be pruned / factored out
def populate_switch(switch):
    try:
        switch.populate_lite()
    except:
        pass


def test_wb_switches():
    global wb
    global switches
    global pool
    global rslts
    wb = WorkbookWrapper('bia-netw.xlsx')
    switches = [switch for switch in wb.switches_from_rows() if switch is not None]
    pool = multiprocessing.pool.ThreadPool(processes=32)
    start_time = time.time()
    rslts = pool.map_async(populate_switch, switches)
    increment_table = {100: 5, 50: 3, 25: 1, 10: 0.5}
    remaining_q = []
    increment = 5
    while True:
        remaining_switches = [switch.ip for switch in switches if switch.state == 'UNK']
        remaining = len(remaining_switches)
        if remaining == 0:
            return
        seconds = time.time() - start_time
        print('{0} remaining after {1} seconds'.format(remaining, seconds))
        for key in sorted(increment_table.keys()):
            if remaining >= key:
                increment = increment_table[key]
            else:
                break

        if remaining in remaining_q:  # at least one nonproductive cycle
            if len(remaining_q) == 4:
                print('Remaining switches:')
                pprint(remaining_switches)
        else:
            remaining_q = []

        remaining_q.append(remaining)

        time.sleep(increment)

    pool.close()
    pool.join()

def test_wb_switches_simple(max_runs):
    global wb
    global switches
    global pool
    global rslts
    wb = WorkbookWrapper('bia-netw.xlsx')
    switches = [switch for switch in wb.switches_from_rows() if switch is not None]

    rslts = map(populate_switch, switches[:max_runs])
    return rslts